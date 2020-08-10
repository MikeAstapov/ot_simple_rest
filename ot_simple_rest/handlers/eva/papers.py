import os
import io
import jwt
import uuid
import json
import docx
import openpyxl
import tempfile
import base64
import zipfile
from datetime import datetime, timedelta
import logging
from preview_generator.manager import PreviewManager

import tornado.web
import tornado.httputil

from handlers.eva.base import BaseHandler

__author__ = "Fedor Metelkin"
__copyright__ = "Copyright 2020, ISG Neuro"
__credits__ = []
__license__ = ""
__version__ = "0.0.1"
__maintainer__ = "Andrey Starchenkov"
__email__ = "astarchenkov@isgneuro.com"
__status__ = "Develop"


class PaperLoadHandler(BaseHandler):   # метод отвечающяя за загрузку файл в папку
    def initialize(self, **kwargs):  # инициализируем переменные которые придут при вызове этого метода
        super().initialize(kwargs['db_conn_pool'])
        self.static_conf = kwargs['static_conf']
        self.logger = logging.getLogger('osr')

    async def prepare(self): #  метод без которого не будет работать класс, в нем мы проверяем куку и разрешение на дальнейшую работу
        client_token = self.get_cookie('eva_token')
        if client_token:
            self.token = client_token
            try:
                token_data = self.decode_token(client_token)
                user_id = token_data['user_id']
                self.permissions = self.db.get_permissions_data(user_id=user_id,
                                                                names_only=True)
            except (jwt.ExpiredSignatureError, jwt.DecodeError):
                pass
            else:
                self.current_user = user_id

        if not self.current_user:
            raise tornado.web.HTTPError(401, "unauthorized")

    async def post(self): # метод который положит файл в папку
        body = self.request.body # получаем данные с фронта
        args = {}
        files = {}
        result = {'status': 'success'}
        tornado.httputil.parse_body_arguments(self.request.headers['Content-Type'], body, args, files) # парсим и получаем данные в нужном нам виде
        reports_path = os.path.join(self.static_conf['static_path'], 'reports')  # путь куда будем сохранять файл

        try:
            _file = files['file'][0] # тут из всех переданных данных забираем собственно файл
            saving_full_path = os.path.join(reports_path, _file['filename']) # тут к пути еще добовляем имя файла
            with open(saving_full_path, 'wb') as f: # то открывает его (и создает видимо)
                try:
                    f.write(_file['body'])  # и записывает в него данные с фронта
                except Exception:
                    result['status'] = 'failed'
                    reuslt['description'] = 'fail is not write' 
        except AttributeError:
            result['status'] = 'failed'
            reuslt['description'] = 'fail is not send'
            
        # if not os.path.exists(saving_full_path): # тут не до конца понимаю как работает но в общем проверят есть ли уже такой файл и если нет
        self.write(result) # передаем успешное выполнение запроса

class PapersHandler(BaseHandler):  # метод возвращающий все файлы в папке на фронт
    def initialize(self, **kwargs):   # инициализируем переменные которые придут при вызове этого метода
        super().initialize(kwargs['db_conn_pool'])
        self.static_conf = kwargs['static_conf']
        self.logger = logging.getLogger('osr')

    async def prepare(self): #  метод без которого не будет работать класс, в нем мы проверяем куку и разрешение на дальнейшую работу
        client_token = self.get_cookie('eva_token')
        if client_token:
            self.token = client_token
            try:
                token_data = self.decode_token(client_token)
                user_id = token_data['user_id']
                self.permissions = self.db.get_permissions_data(user_id=user_id,
                                                                names_only=True)
            except (jwt.ExpiredSignatureError, jwt.DecodeError):
                pass
            else:
                self.current_user = user_id

        if not self.current_user:
            raise tornado.web.HTTPError(401, "unauthorized")

    async def get(self): # метод который вернет все файлы в папке
      reports_path = self.static_conf['static_path'] + 'reports'  # путь до нужной папки
      files = []  # переменная в которой будет храниться список файлов
      for file in os.listdir(reports_path): # забираем список всех файлов и каталогов из нужной папки
        if os.path.isfile(os.path.join(reports_path, file)): # проверяем если это файл, а не каталог
          files.append(file) # то заносим этот файл в подготовленный массив
      # listOfFiles = [f for f in os.listdir(reports_path) if os.path.isfile(f)] ну почти.. прием хороший, но дебажить его не возможно
      # print(files)
      
      self.write({'files':files,'status': 'success'}) # возвращаем список файлов и сообщение что успешно все прошло 


class PaperHandler(BaseHandler): # метод который изменит файл с фротна и вернет ссылку на новый
    def initialize(self, **kwargs): # инициализируем переменные которые придут при вызове этого метода
        super().initialize(kwargs['db_conn_pool'])
        self.static_conf = kwargs['static_conf']
        self.data_path = kwargs['mem_conf']['path']
        self.logger = logging.getLogger('osr')
        self.static_dir_name = 'storage'

    async def prepare(self):   # инициализируем переменные которые придут при вызове этого метода
        client_token = self.get_cookie('eva_token')
        if client_token:
            self.token = client_token
            try:
                token_data = self.decode_token(client_token)
                user_id = token_data['user_id']
                self.permissions = self.db.get_permissions_data(user_id=user_id,
                                                                names_only=True)
            except (jwt.ExpiredSignatureError, jwt.DecodeError):
                pass
            else:
                self.current_user = user_id

        if not self.current_user:
            raise tornado.web.HTTPError(401, "unauthorized")
        

    async def post(self): # метод который изменит файл на основе данных с фронта и отдаст ссылку на новый
        body = self.request.body # получаем данные с фронта
        args = {}
        tornado.httputil.parse_body_arguments(self.request.headers['Content-Type'], body, args, {}) # парсим и получаем данные в нужном нам виде
        file_name = args['file'][0].decode('utf-8')  # получаем имя нужного файла после раскадировки
        reports_path = os.path.join(self.static_conf['static_path'],  'reports')   # путь до папки откуда брать файлы
        try:  # попытаемся получить готовые данные с фронта
          data = args['data'][0].decode('utf-8') # если получить удалось
          data = {"status": "success", "data": json.loads(data)} # то подготавливаем дату в нужный нам вид и записываем туды данные
        except KeyError: # если такого ключа нет, значит нам эти данные нужно получить самим
          cid = args['cid'][0].decode('utf-8')  # получаем cid запроса
          data = self.get_data(cid)  # вызываем метод для получения данных
          for i, json_data in enumerate(data['data']): # так же нам надо перевести строки json в dist  поэтому пробегаемся по всем данным
            data['data'][i] = json.loads(json_data)  # и распаршиваем json данные в dict
        except:  # если произошла люябая другая ошибка
          self.write({'description':'data is not available','status': 'failed'}) #  выведем сообщение что данные поулчить не удалось
          
        if data['status'] == 'failed':
          self.write({'description':'cache is cleared and search is gone','status': 'failed'})
        else:
          data = data['data']
          full_path =  os.path.join(reports_path,  file_name)  # полный путь уже с именем нужного файла
          about_file =  file_name.split('.') # получаем массив в котором первое значение имя файла, а второе его разрешение

          if about_file[1] == 'xlsx': # если файл с разрешением xlsx 
            file_res = self.work_xlsx(full_path,data,about_file[0]) # то вызываем метод который обработает xlsx
            result = {'file':file_res['link'],'html':file_res['html'],'names':file_res['names'],'status': 'success'} # и заносим результат в переменную
          elif about_file[1] == 'docx': # если расширение docx
            file_res =self.work_docx(full_path,data,about_file[0]) # то вызываем метод который обработает docx
            result = {'file':file_res['link'],'html':file_res['html'],'names':file_res['names'],'status': 'success'} # и заносим результат в переменную
          else: # если файл не поддерживается
            result = {'description':'file is not supported','status': 'failed'} # то вернем статус об ошибке

          self.write(result) # вернем результат выполнения метода

    def work_docx(self,path,data,name_file): # метод для работы с xlsx файлами
      # TODO По-хорошему, надо вынести в отдельный модуль. Преобразование файлов не должно быть частью хэндлера.
      result = ''
      files= []
      html = []
      reports_path = os.path.join(self.static_conf['static_path'],  'reports/changed')  # задаем правлиьный путь для измененных файлов

      for i, part_data in enumerate(data):
       
        doc = docx.Document(path)

        for paragraph in doc.paragraphs:  
          for key in part_data.keys(): # пробегаемся по словарю данных с фронта
            if  paragraph.text.find('$'+key+'$') != -1: # а затем проверяем есть ли в этой ячейке ключ словаря
              paragraph.text = paragraph.text.replace('$'+key+'$', part_data[key])

        if (len(data) > 1):   # если несколько строк данных
          file_full_name = f"{name_file}_{datetime.strftime(datetime.now()+ timedelta(seconds=i), '%Y%m%d%H%M%S')}.docx" # то создаем несоклько файлов но каждому следующему увеличиваем время на секунду
        else: # если строка всего одна
          file_full_name = f"{name_file}_{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.docx" # то просто задаем ей имя исходя из времени создания

        files.append(file_full_name) # уже полный путь с названием файла
        doc.save(os.path.join(reports_path, file_full_name))  # сохраняем измененный файл в папку

       
        with tempfile.TemporaryDirectory() as directory:
          preview_path = os.path.join(reports_path, file_full_name)

          manager = PreviewManager(directory, create_folder= True)
          path_to_preview_image = manager.get_jpeg_preview(preview_path,height=1080,width=1924)

          with open(path_to_preview_image, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read())
            html.append(encoded_string.decode('utf-8'))   


      if len(files) > 1: #  если у нас несколько файлов

        result = {"link" : 'reports/changed/'+self.to_archive(name_file,files,reports_path), "html" : html,"names": files} # задаем путь до архива, вызвав метод для создания архивов

      else:  # если файл только один
        result = {"link" : f"reports/changed/{file_full_name}", "html" : html, "names": files}  # просто указываем путь до архива

      return result # возвращаем ссылку на измененный файл 

    def work_xlsx(self,path,data,name_file): # метод для работы с xlsx файлами
      # TODO По-хорошему, надо вынести в отдельный модуль Преобразование файлов не должно быть частью хэндлера.

      files = []
      reports_path = os.path.join(self.static_conf['static_path'],  'reports/changed')  # задаем правлиьный путь для измененных файлов
      result = {}
      html = []

     
      for i, part_data in enumerate(data):


        wb = openpyxl.load_workbook(path) # открываем файл
        sheet = wb.active  # выбираем активный лист
  
        for rownum in range(sheet.max_row): # пробегаемся по всем строкам 
          for columnnum in range(sheet.max_column): #  и в каждой строке по всем столбцам
            cell = sheet.cell(rownum + 1, columnnum + 1).value #  запоминаем занчение в текущей ячейки
            for key in part_data.keys(): # пробегаемся по словарю данных с фронта
              if cell is not None and type(cell) is str: # првоеряем не пустая ли ячейка и что ячейка строка 
                if  cell.find('$'+key+'$') != -1: # а затем проверяем есть ли в этой ячейке ключ словаря
                  cell = cell.replace('$'+key+'$', part_data[key])  # то заменяем значение ячейке на значение из данных
                  sheet.cell(rownum + 1, columnnum + 1).value = cell # Записываем измененую ячейку в файл
        
        if (len(data) > 1): # если несколько строк данных
          file_full_name = f"{name_file}_{datetime.strftime(datetime.now()+ timedelta(seconds=i), '%Y%m%d%H%M%S')}.xlsx" # то создаем несоклько файлов но каждому следующему увеличиваем время на секунду
        else: # если строка всего одна
          file_full_name = f"{name_file}_{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.xlsx"  # то просто задаем ей имя исходя из времени создания

        files.append(file_full_name) # уже полный путь с названием файла


        wb.save(os.path.join(reports_path, file_full_name)) # сохраняем измененный файл в папку

        with tempfile.TemporaryDirectory() as directory:
          preview_path = os.path.join(reports_path, file_full_name)

          manager = PreviewManager(directory, create_folder= True)
          path_to_preview_image = manager.get_jpeg_preview(preview_path,height=1080,width=1924)

          with open(path_to_preview_image, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read())
            html.append(encoded_string.decode('utf-8')) 

        # out_stream = xlsx2html(os.path.join(reports_path, file_full_name))
        # out_stream.seek(0)
        # html.append(out_stream.read())

      if len(files) > 1: #  если у нас несколько файлов

        result = {"link" : 'reports/changed/'+self.to_archive(name_file,files,reports_path), "html" : html, "names": files} # задаем путь до архива, вызвав метод для создания архивов

      else:  # если файл только один
        result = {"link" : f"reports/changed/{file_full_name}", "html" : html, "names": files} # просто указываем путь до архива

  

      return result # возвращаем ссылку на измененный файл 


    def to_archive(self,name_file,files,reports_path):
      
      with tempfile.TemporaryDirectory() as directory: # создаем временную папку ## ну ты и лентяй

        archive_name = f"{name_file}_{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}_archive.zip"
        archive_path = os.path.join(directory, archive_name)  # задаем путь до архива во временной папке

        archive=zipfile.ZipFile(archive_path,'w') #создаем архив
        for name in files:  # пробегаемся по всем файлам
          os.rename(os.path.join(reports_path, name), os.path.join(directory, name)) # перемещаем созданные файлы во временную папку
          archive.write(os.path.join(directory, name), arcname=name) #добавляем файл в архив
        archive.close() #закрываем архив
        os.rename(os.path.join(directory, archive_name), os.path.join(reports_path, archive_name)) # переносим архив в папку с изменнными файлами

      return archive_name

    def get_data(self, cid): # метод для поулчения данных запроса
      body = []
      try:   # попробуем
        path_to_cache_dir = os.path.join(self.data_path, f'search_{cid}.cache/data') #  получить путь до файла, и если удалось
        file_names = [file_name for file_name in os.listdir(path_to_cache_dir) if file_name[-5:] == '.json'] # то берем все файлы с разрешением json
        for file_name in file_names:
            with open(os.path.join(path_to_cache_dir, file_name)) as fr: # открываем его для прочтения
                for i, line in enumerate(fr):
                  if line != '' and i < 100:
                    body.append(line.replace('\n', ''))
                    i = i + 1
        result = {"status": "success","data": body}  # отдаем успешный статус и наши данные
      except: # если не получилось достучаться до файла
        result = {"status": "failed"}  # то скорее всего время жизни кэша истекло и его больше нет, возвращаем статус failed

      return result
      



