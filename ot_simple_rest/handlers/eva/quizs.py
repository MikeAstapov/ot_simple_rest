import os
import io
import json
import tarfile
import tempfile
import uuid
from datetime import datetime

import tornado.web
import tornado.httputil
from openpyxl import Workbook

from handlers.eva.base import BaseHandler


class QuizsHandler(BaseHandler):
    """
    That handler allows to get list of quizs with offset & limit params for pagination.
    """

    async def get(self):
        _offset = self.get_argument('offset', 0)
        _limit = self.get_argument('limit', 10)

        quizs = self.db.get_quizs(limit=_limit, offset=_offset)
        self.write({'data': quizs, 'count': self.db.get_quizs_count()})


class QuizHandler(BaseHandler):
    """
    There is four methods for four actions with quiz objects.
    - get:      returns quiz data by 'id' param;
    - post:     creates new quiz object in DB with data from json body;
    - put:      edit existing quiz object in DB with data from json body;
    - delete:   delete existing quiz object from DB by 'id' param;
    """

    async def get(self):
        quiz_id = self.get_argument('id', None)
        if not quiz_id:
            raise tornado.web.HTTPError(400, "param 'id' is needed")
        try:
            quiz = self.db.get_quiz(quiz_id=quiz_id)
        except Exception as err:
            raise tornado.web.HTTPError(400, str(err))
        self.write({'data': quiz})

    async def post(self):
        quiz_name = self.data.get('name', None)
        questions = self.data.get('questions', None)
        if None in [quiz_name, questions]:
            raise tornado.web.HTTPError(400, "params 'name' and 'questions' is needed")
        try:
            quiz_id = self.db.add_quiz(name=quiz_name,
                                       questions=questions)
        except Exception as err:
            raise tornado.web.HTTPError(409, str(err))
        self.write({'id': quiz_id})

    async def put(self):
        quiz_id = self.data.get('id', None)
        quiz_name = self.data.get('name', None)
        questions = self.data.get('questions', None)
        if not quiz_id:
            raise tornado.web.HTTPError(400, "param 'id' is needed")
        try:
            quiz_id = self.db.update_quiz(quiz_id=quiz_id,
                                          name=quiz_name,
                                          questions=questions)
        except Exception as err:
            raise tornado.web.HTTPError(409, str(err))
        self.write({'id': quiz_id})

    async def delete(self):
        quiz_id = self.get_argument('id', None)
        if not quiz_id:
            raise tornado.web.HTTPError(400, "params 'id' is needed")
        quiz_id = self.db.delete_quiz(quiz_id=quiz_id)
        self.write({'id': quiz_id})


class QuizFilledHandler(BaseHandler):
    """
    Handling actions with filled quiz object.
    It's allows two actions:
    - get:      gets filled quiz object from DB by 'id' param and limit/offset params for pagination;
    - post:     adds new fille quiz object to DB with data from json body;
    """

    async def get(self):
        quiz_type_id = self.get_argument('id', None)
        offset = self.get_argument('offset', 0)
        limit = self.get_argument('limit', 10)

        try:
            filled_quizs = self.db.get_filled_quiz(quiz_id=quiz_type_id,
                                                   offset=offset,
                                                   limit=limit)
        except Exception as err:
            raise tornado.web.HTTPError(409, str(err))
        count = self.db.get_filled_quizs_count(quiz_type_id) if quiz_type_id else len(filled_quizs)
        self.write({'data': filled_quizs, 'count': count})

    async def post(self):
        filled_ids = list()
        for quiz in self.data:
            quiz_type_id = quiz.get('id', None)
            questions = quiz.get('questions', None)

            if None in [quiz_type_id, questions]:
                raise tornado.web.HTTPError(400, "params 'id', and 'questions' is needed")
            try:
                self.db.save_filled_quiz(user_id=self.current_user,
                                         quiz_id=quiz_type_id,
                                         questions=questions)
                filled_ids.append(quiz_type_id)
            except Exception as err:
                raise tornado.web.HTTPError(409, str(err))
        self.write({'ids': filled_ids})


class QuizQuestionsHandler(BaseHandler):
    """
    If is need a question list for one or more quiz, this handler for it.
    Input param is 'ids', like '?ids=1,2,3'.
    """

    async def get(self):
        quiz_ids = self.get_argument('ids', None)
        if not quiz_ids:
            raise tornado.web.HTTPError(400, "params 'ids' is needed")
        quiz_ids = quiz_ids.split(',')
        quiz_ids = [int(i) for i in quiz_ids if i]
        try:
            questions = self.db.get_quiz_questions(quiz_ids=quiz_ids)
        except Exception as err:
            raise tornado.web.HTTPError(409, str(err))
        self.write({'data': questions})


class FilledQuizExportHandler(BaseHandler):
    """
    This handler allows export filled quiz object into '.xlsx' format file.
    Input param is 'id' which quiz.id in DB.
    """

    def initialize(self, **kwargs):
        super().initialize(kwargs['db_conn_pool'])
        self.static_conf = kwargs['static_conf']

    async def get(self):
        quiz_id = self.get_argument('id', None)
        if not quiz_id:
            raise tornado.web.HTTPError(400, "param 'id' is needed")
        quiz_id = int(quiz_id)
        try:
            quiz_data = self.db.get_filled_quiz(quiz_id=quiz_id, current=True)
            quiz_data = quiz_data[0] if quiz_data else None
            if not quiz_data:
                raise tornado.web.HTTPError(404, f'No quiz with id={quiz_id}')
        except Exception as err:
            raise tornado.web.HTTPError(409, str(err))

        questions = quiz_data['questions']
        quiz_name = quiz_data['name'].replace('«', '"').replace('»', '"')
        quiz_name = quiz_name.replace(' ', '_')
        filled = quiz_data['fill_date'].replace(' ', '').replace(':', '')

        filename = f'{quiz_name}_{filled}.xlsx'
        filepath = os.path.join(self.static_conf['static_path'], 'xlsx', filename)
        if os.path.exists(filepath):
            return self.write(f'/static/xlsx/{filename}')

        wb = Workbook()
        ws = wb.active
        ws.title = 'чек-лист'
        cell_range = ws['A1':'F1']
        col_range = ['№', 'Вопрос', 'Ответ', 'Пояснение', 'Ключевой', 'Метка']
        for c, v in zip(*cell_range, col_range):
            c.value = v

        for i, q in enumerate(questions, 2):
            cell_range = ws[f'A{i}':f'F{i}']
            col_range = [q['sid'], q['text'], q['answer']['value'], q['answer']['description'],
                         q['is_sign'], q['label']]
            for c, v in zip(*cell_range, col_range):
                c.value = v

        wb.save(filepath)
        self.write(f'/static/xlsx/{filename}')


class QuizExportJsonHandler(BaseHandler):
    """
    There is method for export one or more quiz object in '.json' format files.
    Json files returns in 'tar.gz' package with uuid-name.
    """

    def initialize(self, **kwargs):
        super().initialize(kwargs['db_conn_pool'])
        self.static_conf = kwargs['static_conf']

    async def get(self):
        quiz_ids = self.get_argument('ids', None)
        if not quiz_ids:
            raise tornado.web.HTTPError(400, "param 'ids' is needed")
        quiz_ids = quiz_ids.split(',')
        quiz_ids = [int(_) for _ in quiz_ids]

        with tempfile.TemporaryDirectory() as tmp_dir:
            archive_name = f"{datetime.strftime(datetime.now(), '%Y%m%d%H%M%S')}.eva.quiz"
            _dirname = str(uuid.uuid4())
            _base_path = os.path.join(self.static_conf['static_path'], 'storage', _dirname)
            if not os.path.exists(_base_path):
                os.makedirs(_base_path)

            archive_path = os.path.join(_base_path, archive_name)
            archive = tarfile.open(archive_path, mode='x:gz')

            for qid in quiz_ids:
                try:
                    quiz_data = self.db.get_quiz(quiz_id=qid)
                    if not quiz_data:
                        raise tornado.web.HTTPError(404, f'No quiz with id={qid}')
                except Exception as err:
                    raise tornado.web.HTTPError(409, str(err))

                filename = f'{qid}.json'
                filepath = os.path.join(tmp_dir, filename)

                if not os.path.exists(filepath):
                    with open(filepath, 'w+') as f:
                        f.write(json.dumps(quiz_data, ensure_ascii=False))

                archive.add(filepath, filename)
            archive.close()
        self.write(f'storage/{_dirname}/{archive_name}')


class QuizImportJsonHandler(BaseHandler):
    """
    That handler allows to import quizs, exported with QuizExportJsonHandler.
    Or you can put your own 'tar.gz' file with inner quizs json files.
    """

    async def prepare(self):
        client_token = self.get_cookie('eva_token')
        if client_token:
            self.token = client_token
            try:
                token_data = self.decode_token(client_token)
                user_id = token_data['user_id']
                self.permissions = self.db.get_permissions_data(user_id=user_id,
                                                                names_only=True)
            except (jwt.ExpiredSignatureError, jwt.DecodeError):
                pass
            else:
                self.current_user = user_id

        if not self.current_user:
            raise tornado.web.HTTPError(401, "unauthorized")

    async def post(self):
        body = self.request.body
        args = {}
        files = {}
        tornado.httputil.parse_body_arguments(self.request.headers['Content-Type'], body, args, files)
        if not files or not files.get('file'):
            return self.write({'status': 'no file in payload'})
        tar_file = files['file'][0]

        # wraps bytes to work with it like with file
        file_like_object = io.BytesIO(tar_file['body'])
        with tarfile.open(mode='r:gz', fileobj=file_like_object) as tar:
            for quiz in tar.getmembers():
                quiz_data = tar.extractfile(quiz)
                quiz_data = json.loads(quiz_data.read())
                quiz_name = quiz_data.get('name', None)
                questions = quiz_data.get('questions', None)
                if None in [quiz_name, questions]:
                    raise tornado.web.HTTPError(400, "params 'name' and 'questions' is needed")
                try:
                    self.db.add_quiz(name=quiz_name,
                                     questions=questions)
                except Exception as err:
                    raise tornado.web.HTTPError(409, str(err))
            self.write({'status': 'success'})


class CatalogsListHandler(BaseHandler):
    def get(self):
        _offset = self.get_argument('offset', 0)
        _limit = self.get_argument('limit', 10)

        catalogs = self.db.get_catalogs_data(limit=_limit, offset=_offset)
        self.write({'data': catalogs, 'count': self.db.get_catalogs_count()})


class CatalogHandler(BaseHandler):
    async def get(self):
        catalog_id = self.get_argument('id', None)
        if not catalog_id:
            raise tornado.web.HTTPError(400, "param 'id' is needed")
        try:
            catalog = self.db.get_catalog(catalog_id=catalog_id)
        except Exception as err:
            raise tornado.web.HTTPError(400, str(err))
        self.write({'data': catalog})

    async def post(self):
        catalog_name = self.data.get('name', None)
        content = self.data.get('content', None)
        if None in [catalog_name, content]:
            raise tornado.web.HTTPError(400, "params 'name' and 'content' is needed")
        try:
            catalog_id = self.db.add_catalog(name=catalog_name,
                                             content=content)
        except Exception as err:
            raise tornado.web.HTTPError(409, str(err))
        self.write({'id': catalog_id})

    async def put(self):
        catalog_id = self.data.get('id', None)
        catalog_name = self.data.get('name', None)
        content = self.data.get('content', None)
        if not catalog_id:
            raise tornado.web.HTTPError(400, "param 'id' is needed")
        try:
            catalog_id = self.db.update_catalog(catalog_id=catalog_id,
                                                name=catalog_name,
                                                content=content)
        except Exception as err:
            raise tornado.web.HTTPError(409, str(err))
        self.write({'id': catalog_id})

    async def delete(self):
        catalog_id = self.get_argument('id', None)
        if not catalog_id:
            raise tornado.web.HTTPError(400, "params 'id' is needed")
        catalog_id = self.db.delete_catalog(catalog_id=catalog_id)
        self.write({'id': catalog_id})

